# -*- coding: utf-8 -*-
"""
民宿房间管理系统 - 主应用入口
基于 Flask，适用于 iStoreOS (OpenWrt) 环境
数据存储采用 JSON 文件，兼具可读性和可迁移性
"""

import os
import sys
import io
from datetime import datetime, timedelta

from flask import (Flask, render_template, request, jsonify,
                   redirect, url_for, send_from_directory, send_file)

# 添加项目根目录到路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from config import *
from utils.data_manager import DataManager
from utils.logger import SystemLogger
from utils.image_handler import ImageHandler

app = Flask(__name__)
app.config['SECRET_KEY'] = SECRET_KEY
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH


# ==================== 辅助函数 ====================

def get_today_str():
    return datetime.now().strftime('%Y-%m-%d')


def parse_date(date_str):
    """解析日期字符串"""
    try:
        return datetime.strptime(date_str, '%Y-%m-%d')
    except (ValueError, TypeError):
        return None


def _get_random_quotes(n=3):
    """从名言文件中随机读取 n 条"""
    try:
        if os.path.exists(QUOTES_FILE):
            with open(QUOTES_FILE, 'r', encoding='utf-8') as f:
                lines = [line.strip() for line in f if line.strip()]
            import random
            return random.sample(lines, min(n, len(lines)))
    except Exception:
        pass
    return ['用心服务每一位客人 🌟']


# ==================== 路由：页面 ====================

@app.route('/')
def index():
    """仪表盘首页"""
    return render_template('index.html', config={
        'homestay_name': HOMESTAY_NAME
    })


@app.route('/room-types')
def room_types_page():
    """房型管理页面"""
    return render_template('room_types.html')


@app.route('/rooms')
def rooms_page():
    """房间管理页面"""
    return render_template('rooms.html')


@app.route('/guests')
def guests_page():
    """客人管理页面"""
    return render_template('guests.html')


@app.route('/bookings')
def bookings_page():
    """预订管理页面"""
    return render_template('bookings.html')


@app.route('/calendar')
def calendar_page():
    """日历视图页面"""
    return render_template('calendar.html')


@app.route('/logs')
def logs_page():
    """日志查看页面"""
    return render_template('logs.html')


@app.route('/diary')
def diary_page():
    """运营日记页面"""
    return render_template('diary.html')


@app.route('/statistics')
def statistics_page():
    """统计页面"""
    return render_template('statistics.html')


@app.route('/booking-sources')
def booking_sources_page():
    """预订来源管理页面"""
    return render_template('booking_sources.html')


@app.route('/holidays')
def holidays_page():
    """节假日管理页面"""
    return render_template('holidays.html')


@app.route('/expenses')
def expenses_page():
    """支出管理页面"""
    return render_template('expenses.html')


@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    """访问上传的文件"""
    return send_from_directory(UPLOAD_DIR, filename)


# ==================== API：统计概览 ====================

@app.route('/api/dashboard')
def api_dashboard():
    """仪表盘统计数据"""
    try:
        today = get_today_str()
        rooms = DataManager.read_json(ROOMS_FILE, [])
        bookings = DataManager.read_json(BOOKINGS_FILE, [])
        guests = DataManager.read_json(GUESTS_FILE, [])

        total_rooms = len(rooms)
        occupied = sum(1 for r in rooms if r.get('status') == 'occupied')
        reserved = sum(1 for r in rooms if r.get('status') == 'reserved')
        available = sum(1 for r in rooms if r.get('status') == 'available')
        cleaning = sum(1 for r in rooms if r.get('status') == 'cleaning')
        maintenance = sum(1 for r in rooms if r.get('status') == 'maintenance')

        # 今日相关预订
        today_checkins = [b for b in bookings if b.get('check_in_date') == today]
        today_checkouts = [b for b in bookings if b.get('check_out_date') == today]
        active_bookings = [b for b in bookings
                           if b.get('status') in ('confirmed', 'checked_in')]

        # 今日收入 = 所有房间今日价格之和
        today_prices = DataManager.read_json(DAILY_PRICES_FILE, [])
        today_income = sum(
            float(p.get('price', 0)) for p in today_prices
            if p.get('date') == today
        )
        # 加上今日额外收入
        extra_incomes = DataManager.read_json(EXTRA_INCOME_FILE, [])
        today_extra = sum(
            float(e.get('amount', 0)) for e in extra_incomes
            if e.get('date') == today
        )
        today_income += today_extra

        # 本月收入 = 所有房间本月每日价格之和 + 本月额外收入
        current_month = datetime.now().strftime('%Y-%m')
        month_prices_sum = sum(
            float(p.get('price', 0)) for p in today_prices
            if p.get('date', '').startswith(current_month)
        )
        month_extra = sum(
            float(e.get('amount', 0)) for e in extra_incomes
            if e.get('date', '').startswith(current_month)
        )
        month_income = month_prices_sum + month_extra

        # 入住率
        occupancy_rate = round((occupied / total_rooms * 100), 1) if total_rooms > 0 else 0

        # 最近预订
        bookings.sort(key=lambda x: x.get('created_at', ''), reverse=True)
        recent_bookings = bookings[:5]

        # 最近日志
        logs_data = SystemLogger.get_logs(page=1, page_size=10)

        # 高考倒计时
        gaokao_date = datetime(2027, 6, 7)
        countdown_days = (gaokao_date - datetime.now()).days

        # 每日名言（随机3条）
        quotes = _get_random_quotes(3)

        return jsonify({
            'success': True,
            'data': {
                'total_rooms': total_rooms,
                'occupied': occupied,
                'reserved': reserved,
                'available': available,
                'cleaning': cleaning,
                'maintenance': maintenance,
                'total_guests': len(guests),
                'today_checkins': len(today_checkins),
                'today_checkouts': len(today_checkouts),
                'active_bookings': len(active_bookings),
                'today_income': round(today_income, 2),
                'month_income': round(month_income, 2),
                'month_prices_sum': round(month_prices_sum, 2),
                'month_extra': round(month_extra, 2),
                'occupancy_rate': occupancy_rate,
                'recent_bookings': recent_bookings,
                'recent_logs': logs_data['items'],
                'countdown_days': countdown_days,
                'quotes': quotes
            }
        })
    except Exception as e:
        SystemLogger.error('获取仪表盘数据失败', str(e))
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== API：房型管理 ====================

@app.route('/api/room-types', methods=['GET', 'POST'])
def api_room_types():
    if request.method == 'GET':
        # 查询房型列表
        keyword = request.args.get('keyword', '')
        page = int(request.args.get('page', 1))
        types = DataManager.read_json(ROOM_TYPES_FILE, [])

        if keyword:
            types = DataManager.search(types, keyword,
                                       ['name', 'description', 'amenities'])

        result = DataManager.query(types, sort_key='id', page=page, page_size=PAGE_SIZE)
        return jsonify({'success': True, 'data': result})

    elif request.method == 'POST':
        # 新增房型
        try:
            data = request.get_json()
            types = DataManager.read_json(ROOM_TYPES_FILE, [])

            new_type = {
                'id': DataManager.generate_id(types),
                'name': data.get('name', '').strip(),
                'description': data.get('description', '').strip(),
                'base_price': float(data.get('base_price', 0)),
                'weekend_price': float(data.get('weekend_price', 0)) or None,
                'holiday_price': float(data.get('holiday_price', 0)) or None,
                'capacity': int(data.get('capacity', 1)),
                'area': data.get('area', '').strip(),
                'bed_type': data.get('bed_type', '').strip(),
                'amenities': data.get('amenities', '').strip(),
                'images': data.get('images', []),
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            types.append(new_type)
            DataManager.write_json(ROOM_TYPES_FILE, types)
            SystemLogger.info('新增房型', f'房型: {new_type["name"]}')
            return jsonify({'success': True, 'data': new_type})
        except Exception as e:
            SystemLogger.error('新增房型失败', str(e))
            return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/room-types/<int:type_id>', methods=['GET', 'PUT', 'DELETE'])
def api_room_type_detail(type_id):
    types = DataManager.read_json(ROOM_TYPES_FILE, [])

    if request.method == 'GET':
        item = DataManager.get_by_id(types, type_id)
        if item:
            return jsonify({'success': True, 'data': item})
        return jsonify({'success': False, 'error': '房型不存在'}), 404

    elif request.method == 'PUT':
        try:
            data = request.get_json()
            updates = {
                'name': data.get('name', '').strip(),
                'description': data.get('description', '').strip(),
                'base_price': float(data.get('base_price', 0)),
                'weekend_price': float(data.get('weekend_price', 0)) or None,
                'holiday_price': float(data.get('holiday_price', 0)) or None,
                'capacity': int(data.get('capacity', 1)),
                'area': data.get('area', '').strip(),
                'bed_type': data.get('bed_type', '').strip(),
                'amenities': data.get('amenities', '').strip(),
                'images': data.get('images', []),
            }
            if DataManager.update_by_id(types, type_id, updates):
                DataManager.write_json(ROOM_TYPES_FILE, types)
                SystemLogger.info('更新房型', f'房型ID: {type_id}')
                return jsonify({'success': True})
            return jsonify({'success': False, 'error': '房型不存在'}), 404
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400

    elif request.method == 'DELETE':
        # 检查是否有关联房间
        rooms = DataManager.read_json(ROOMS_FILE, [])
        linked = [r for r in rooms if r.get('room_type_id') == type_id]
        if linked:
            room_nums = ', '.join(r['room_number'] for r in linked)
            return jsonify({
                'success': False,
                'error': f'该房型下还有房间: {room_nums}，请先删除关联房间'
            }), 400

        if DataManager.delete_by_id(types, type_id):
            DataManager.write_json(ROOM_TYPES_FILE, types)
            SystemLogger.info('删除房型', f'房型ID: {type_id}')
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '房型不存在'}), 404


# ==================== API：房间管理 ====================

@app.route('/api/rooms', methods=['GET', 'POST'])
def api_rooms():
    if request.method == 'GET':
        keyword = request.args.get('keyword', '')
        status = request.args.get('status', '')
        room_type_id = request.args.get('room_type_id', '')
        floor = request.args.get('floor', '')
        page = int(request.args.get('page', 1))

        rooms = DataManager.read_json(ROOMS_FILE, [])
        types = DataManager.read_json(ROOM_TYPES_FILE, [])

        if keyword:
            rooms = DataManager.search(rooms, keyword,
                                       ['room_number', 'notes'])
        if status:
            rooms = [r for r in rooms if r.get('status') == status]
        if room_type_id:
            rooms = [r for r in rooms
                     if str(r.get('room_type_id')) == room_type_id]
        if floor:
            rooms = [r for r in rooms if str(r.get('floor')) == floor]

        result = DataManager.query(rooms, sort_key='sort_order',
                                   page=page, page_size=PAGE_SIZE)

        # 关联房型名称
        type_map = {t['id']: t['name'] for t in types}
        for item in result['items']:
            item['room_type_name'] = type_map.get(item.get('room_type_id'), '未知')

        return jsonify({
            'success': True,
            'data': result,
            'room_types': types,
            'status_options': ROOM_STATUS
        })

    elif request.method == 'POST':
        try:
            data = request.get_json()
            rooms = DataManager.read_json(ROOMS_FILE, [])

            # 检查房间号是否重复
            room_number = data.get('room_number', '').strip()
            if any(r.get('room_number') == room_number for r in rooms):
                return jsonify({'success': False, 'error': '房间号已存在'}), 400

            new_room = {
                'id': DataManager.generate_id(rooms),
                'room_number': room_number,
                'room_type_id': int(data.get('room_type_id', 0)),
                'floor': data.get('floor', '').strip(),
                'status': data.get('status', 'available'),
                'notes': data.get('notes', '').strip(),
                'sort_order': len(rooms),
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            rooms.append(new_room)
            DataManager.write_json(ROOMS_FILE, rooms)
            SystemLogger.info('新增房间', f'房间号: {room_number}')
            return jsonify({'success': True, 'data': new_room})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/rooms/<int:room_id>', methods=['GET', 'PUT', 'DELETE'])
def api_room_detail(room_id):
    rooms = DataManager.read_json(ROOMS_FILE, [])

    if request.method == 'GET':
        item = DataManager.get_by_id(rooms, room_id)
        if item:
            return jsonify({'success': True, 'data': item})
        return jsonify({'success': False, 'error': '房间不存在'}), 404

    elif request.method == 'PUT':
        try:
            data = request.get_json()
            # 检查房间号重复
            room_number = data.get('room_number', '').strip()
            dup = [r for r in rooms
                   if r.get('room_number') == room_number and r.get('id') != room_id]
            if dup:
                return jsonify({'success': False, 'error': '房间号已存在'}), 400

            updates = {
                'room_number': room_number,
                'room_type_id': int(data.get('room_type_id', 0)),
                'floor': data.get('floor', '').strip(),
                'status': data.get('status', 'available'),
                'notes': data.get('notes', '').strip(),
            }
            if DataManager.update_by_id(rooms, room_id, updates):
                DataManager.write_json(ROOMS_FILE, rooms)
                SystemLogger.info('更新房间', f'房间ID: {room_id}')
                return jsonify({'success': True})
            return jsonify({'success': False, 'error': '房间不存在'}), 404
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400

    elif request.method == 'DELETE':
        # 检查是否有活跃预订
        bookings = DataManager.read_json(BOOKINGS_FILE, [])
        active = [b for b in bookings
                  if b.get('room_id') == room_id and
                  b.get('status') in ('pending', 'confirmed', 'checked_in')]
        if active:
            return jsonify({
                'success': False,
                'error': '该房间有活跃预订，请先处理相关预订'
            }), 400

        if DataManager.delete_by_id(rooms, room_id):
            DataManager.write_json(ROOMS_FILE, rooms)
            SystemLogger.info('删除房间', f'房间ID: {room_id}')
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '房间不存在'}), 404


# ==================== API：客人管理 ====================

@app.route('/api/guests', methods=['GET', 'POST'])
def api_guests():
    if request.method == 'GET':
        keyword = request.args.get('keyword', '')
        page = int(request.args.get('page', 1))

        guests = DataManager.read_json(GUESTS_FILE, [])

        if keyword:
            guests = DataManager.search(guests, keyword,
                                        ['name', 'phone', 'id_card', 'nationality'])

        result = DataManager.query(guests, sort_key='id', sort_reverse=True,
                                   page=page, page_size=PAGE_SIZE)

        # 关联最近预订信息
        bookings = DataManager.read_json(BOOKINGS_FILE, [])
        for guest in result['items']:
            guest_bookings = [b for b in bookings
                              if b.get('guest_id') == guest['id']]
            guest['booking_count'] = len(guest_bookings)
            guest['last_booking'] = guest_bookings[-1] if guest_bookings else None

        return jsonify({'success': True, 'data': result})

    elif request.method == 'POST':
        try:
            data = request.get_json()
            guests = DataManager.read_json(GUESTS_FILE, [])

            new_guest = {
                'id': DataManager.generate_id(guests),
                'name': data.get('name', '').strip(),
                'phone': data.get('phone', '').strip(),
                'id_card': data.get('id_card', '').strip(),
                'gender': data.get('gender', ''),
                'nationality': data.get('nationality', '').strip(),
                'birthday': data.get('birthday', ''),
                'email': data.get('email', '').strip(),
                'address': data.get('address', '').strip(),
                'notes': data.get('notes', '').strip(),
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            guests.append(new_guest)
            DataManager.write_json(GUESTS_FILE, guests)
            SystemLogger.info('新增客人', f'客人: {new_guest["name"]}')
            return jsonify({'success': True, 'data': new_guest})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/guests/<int:guest_id>', methods=['GET', 'PUT', 'DELETE'])
def api_guest_detail(guest_id):
    guests = DataManager.read_json(GUESTS_FILE, [])

    if request.method == 'GET':
        item = DataManager.get_by_id(guests, guest_id)
        if item:
            # 获取该客人的预订历史
            bookings = DataManager.read_json(BOOKINGS_FILE, [])
            item['bookings'] = [b for b in bookings if b.get('guest_id') == guest_id]
            item['bookings'].sort(key=lambda x: x.get('check_in_date', ''), reverse=True)
            return jsonify({'success': True, 'data': item})
        return jsonify({'success': False, 'error': '客人不存在'}), 404

    elif request.method == 'PUT':
        try:
            data = request.get_json()
            updates = {
                'name': data.get('name', '').strip(),
                'phone': data.get('phone', '').strip(),
                'id_card': data.get('id_card', '').strip(),
                'gender': data.get('gender', ''),
                'nationality': data.get('nationality', '').strip(),
                'birthday': data.get('birthday', ''),
                'email': data.get('email', '').strip(),
                'address': data.get('address', '').strip(),
                'notes': data.get('notes', '').strip(),
            }
            if DataManager.update_by_id(guests, guest_id, updates):
                DataManager.write_json(GUESTS_FILE, guests)
                SystemLogger.info('更新客人信息', f'客人ID: {guest_id}')
                return jsonify({'success': True})
            return jsonify({'success': False, 'error': '客人不存在'}), 404
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400

    elif request.method == 'DELETE':
        # 检查是否有关联预订
        bookings = DataManager.read_json(BOOKINGS_FILE, [])
        linked = [b for b in bookings if b.get('guest_id') == guest_id]
        if linked:
            return jsonify({
                'success': False,
                'error': '该客人有关联预订记录，请先删除相关预订'
            }), 400

        if DataManager.delete_by_id(guests, guest_id):
            DataManager.write_json(GUESTS_FILE, guests)
            SystemLogger.info('删除客人', f'客人ID: {guest_id}')
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '客人不存在'}), 404


# ==================== API：预订管理 ====================

@app.route('/api/bookings', methods=['GET', 'POST'])
def api_bookings():
    if request.method == 'GET':
        keyword = request.args.get('keyword', '')
        status = request.args.get('status', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        page = int(request.args.get('page', 1))

        bookings = DataManager.read_json(BOOKINGS_FILE, [])
        guests = DataManager.read_json(GUESTS_FILE, [])
        rooms = DataManager.read_json(ROOMS_FILE, [])

        if keyword:
            # 通过客人姓名搜索
            matched_guest_ids = [
                g['id'] for g in guests
                if keyword.lower() in g.get('name', '').lower() or
                   keyword.lower() in g.get('phone', '').lower()
            ]
            # 通过房间号搜索
            matched_room_ids = [
                r['id'] for r in rooms
                if keyword.lower() in r.get('room_number', '').lower()
            ]
            bookings = [
                b for b in bookings
                if b.get('guest_id') in matched_guest_ids or
                   b.get('room_id') in matched_room_ids or
                   keyword.lower() in str(b.get('id', ''))
            ]

        if status:
            bookings = [b for b in bookings if b.get('status') == status]

        if date_from:
            bookings = [b for b in bookings
                        if b.get('check_out_date', '') >= date_from]
        if date_to:
            bookings = [b for b in bookings
                        if b.get('check_in_date', '') <= date_to]

        result = DataManager.query(bookings, sort_key='check_in_date',
                                   sort_reverse=True, page=page,
                                   page_size=PAGE_SIZE)

        # 关联客人和房间信息
        guest_map = {g['id']: g for g in guests}
        room_map = {r['id']: r for r in rooms}
        sources = DataManager.read_json(BOOKING_SOURCES_FILE, [])
        source_map = {s['name']: s for s in sources}
        for item in result['items']:
            g = guest_map.get(item.get('guest_id'), {})
            item['guest_name'] = g.get('name', '未知')
            item['guest_phone'] = g.get('phone', '')
            r = room_map.get(item.get('room_id'), {})
            item['room_number'] = r.get('room_number', '未知')
            src = source_map.get(item.get('source', ''), {})
            item['source_color'] = src.get('color', '#6b7280')

        return jsonify({
            'success': True,
            'data': result,
            'status_options': BOOKING_STATUS,
            'sources': sources
        })

    elif request.method == 'POST':
        try:
            data = request.get_json()
            bookings = DataManager.read_json(BOOKINGS_FILE, [])

            # 检查日期冲突
            room_id = int(data.get('room_id', 0))
            check_in = data.get('check_in_date', '')
            check_out = data.get('check_out_date', '')

            conflicts = [
                b for b in bookings
                if b.get('room_id') == room_id and
                   b.get('status') in ('pending', 'confirmed', 'checked_in') and
                   not (check_out <= b.get('check_in_date', '') or
                        check_in >= b.get('check_out_date', ''))
            ]
            if conflicts:
                return jsonify({
                    'success': False,
                    'error': '该时段房间已被预订，请选择其他日期或房间'
                }), 400

            new_booking = {
                'id': DataManager.generate_id(bookings),
                'room_id': room_id,
                'guest_id': int(data.get('guest_id', 0)),
                'check_in_date': check_in,
                'check_out_date': check_out,
                'guest_count': int(data.get('guest_count', 1)),
                'total_price': float(data.get('total_price', 0)),
                'deposit': float(data.get('deposit', 0)),
                'status': data.get('status', 'pending'),
                'notes': data.get('notes', '').strip(),
                'source': data.get('source', '直接预订'),
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            bookings.append(new_booking)
            DataManager.write_json(BOOKINGS_FILE, bookings)

            # 自动填充每日价格到日历
            _auto_fill_daily_prices(room_id, check_in, check_out,
                                    float(data.get('total_price', 0)))

            # 更新房间状态
            rooms = DataManager.read_json(ROOMS_FILE, [])
            DataManager.update_by_id(rooms, room_id, {'status': 'reserved'})
            DataManager.write_json(ROOMS_FILE, rooms)

            SystemLogger.info('新增预订',
                              f'房间: {room_id}, 入住: {check_in}, 退房: {check_out}')
            return jsonify({'success': True, 'data': new_booking})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/bookings/<int:booking_id>', methods=['GET', 'PUT', 'DELETE'])
def api_booking_detail(booking_id):
    bookings = DataManager.read_json(BOOKINGS_FILE, [])

    if request.method == 'GET':
        item = DataManager.get_by_id(bookings, booking_id)
        if item:
            guests = DataManager.read_json(GUESTS_FILE, [])
            rooms = DataManager.read_json(ROOMS_FILE, [])
            guest = DataManager.get_by_id(guests, item.get('guest_id'))
            room = DataManager.get_by_id(rooms, item.get('room_id'))
            item['guest'] = guest
            item['room'] = room
            return jsonify({'success': True, 'data': item})
        return jsonify({'success': False, 'error': '预订不存在'}), 404

    elif request.method == 'PUT':
        try:
            data = request.get_json()
            updates = {
                'room_id': int(data.get('room_id', 0)),
                'guest_id': int(data.get('guest_id', 0)),
                'check_in_date': data.get('check_in_date', ''),
                'check_out_date': data.get('check_out_date', ''),
                'guest_count': int(data.get('guest_count', 1)),
                'total_price': float(data.get('total_price', 0)),
                'deposit': float(data.get('deposit', 0)),
                'status': data.get('status', 'pending'),
                'notes': data.get('notes', '').strip(),
                'source': data.get('source', '直接预订'),
            }

            old_status = DataManager.get_by_id(bookings, booking_id)
            old_status = old_status.get('status') if old_status else None

            if DataManager.update_by_id(bookings, booking_id, updates):
                DataManager.write_json(BOOKINGS_FILE, bookings)

                # 自动更新每日价格（支持强制覆盖）
                force = request.args.get('force_overwrite') == '1'
                _auto_fill_daily_prices(
                    updates['room_id'], updates['check_in_date'],
                    updates['check_out_date'], updates['total_price'],
                    force=force
                )

                # 同步更新房间状态
                new_status = updates['status']
                room_id = updates['room_id']
                rooms = DataManager.read_json(ROOMS_FILE, [])

                status_map = {
                    'checked_in': 'occupied',
                    'checked_out': 'cleaning',
                    'cancelled': 'available',
                    'no_show': 'available',
                }
                if new_status in status_map:
                    DataManager.update_by_id(rooms, room_id,
                                             {'status': status_map[new_status]})
                elif new_status in ('pending', 'confirmed'):
                    DataManager.update_by_id(rooms, room_id,
                                             {'status': 'reserved'})
                DataManager.write_json(ROOMS_FILE, rooms)

                SystemLogger.info('更新预订',
                                  f'预订ID: {booking_id}, 状态: {old_status} -> {new_status}')
                return jsonify({'success': True})
            return jsonify({'success': False, 'error': '预订不存在'}), 404
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400

    elif request.method == 'DELETE':
        booking = DataManager.get_by_id(bookings, booking_id)
        if booking and booking.get('status') in ('checked_in',):
            return jsonify({
                'success': False,
                'error': '已入住的预订不能删除，请先办理退房'
            }), 400

        if DataManager.delete_by_id(bookings, booking_id):
            DataManager.write_json(BOOKINGS_FILE, bookings)
            # 恢复房间状态
            if booking:
                rooms = DataManager.read_json(ROOMS_FILE, [])
                DataManager.update_by_id(rooms, booking['room_id'],
                                         {'status': 'available'})
                DataManager.write_json(ROOMS_FILE, rooms)

                # 是否同时删除每日价格
                if request.args.get('delete_prices') == '1':
                    _delete_daily_prices_range(
                        booking['room_id'],
                        booking.get('check_in_date', ''),
                        booking.get('check_out_date', '')
                    )
            SystemLogger.info('删除预订', f'预订ID: {booking_id}')
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '预订不存在'}), 404


# ==================== API：搜索 ====================

@app.route('/api/search')
def api_search():
    """全局搜索"""
    keyword = request.args.get('q', '').strip()
    if not keyword:
        return jsonify({'success': True, 'data': {
            'room_types': [], 'rooms': [], 'guests': [], 'bookings': []
        }})

    room_types = DataManager.read_json(ROOM_TYPES_FILE, [])
    rooms = DataManager.read_json(ROOMS_FILE, [])
    guests = DataManager.read_json(GUESTS_FILE, [])
    bookings = DataManager.read_json(BOOKINGS_FILE, [])

    # 关联信息
    guest_map = {g['id']: g for g in guests}
    room_map = {r['id']: r for r in rooms}

    results = {
        'room_types': DataManager.search(
            room_types, keyword, ['name', 'description', 'amenities'])[:5],
        'rooms': DataManager.search(
            rooms, keyword, ['room_number', 'floor', 'notes'])[:5],
        'guests': DataManager.search(
            guests, keyword, ['name', 'phone', 'id_card', 'nationality'])[:5],
        'bookings': []
    }

    # 搜索预订（通过房间号和客人姓名）
    matched_room_ids = [r['id'] for r in results['rooms']]
    matched_guest_ids = [g['id'] for g in results['guests']]
    for b in bookings:
        if b.get('room_id') in matched_room_ids or \
           b.get('guest_id') in matched_guest_ids:
            b['room_number'] = room_map.get(b.get('room_id'), {}).get('room_number', '')
            b['guest_name'] = guest_map.get(b.get('guest_id'), {}).get('name', '')
            results['bookings'].append(b)
    results['bookings'] = results['bookings'][:5]

    return jsonify({'success': True, 'data': results})


# ==================== API：统计分析 ====================

@app.route('/api/statistics')
def api_statistics():
    """健壮的统计分析：收支净利润 + 入住率，支持时间范围和房间筛选"""
    mode = request.args.get('mode', 'monthly')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    room_ids_str = request.args.get('room_ids', '')

    daily_prices = DataManager.read_json(DAILY_PRICES_FILE, [])
    extra_incomes = DataManager.read_json(EXTRA_INCOME_FILE, [])
    expenses = DataManager.read_json(EXPENSES_FILE, [])
    rooms = DataManager.read_json(ROOMS_FILE, [])

    # 房间筛选（支持多选，逗号分隔）
    if room_ids_str:
        room_ids = set(room_ids_str.split(','))
        daily_prices = [p for p in daily_prices if str(p.get('room_id')) in room_ids]
        expenses = [e for e in expenses if str(e.get('room_id')) in room_ids]
        rooms = [r for r in rooms if str(r.get('id')) in room_ids]
    room_count = len(rooms)

    # 时间范围
    df = parse_date(date_from) if date_from else None
    dt = parse_date(date_to) if date_to else None

    def _in_range(d):
        if df and d < df.strftime('%Y-%m-%d'): return False
        if dt and d > dt.strftime('%Y-%m-%d'): return False
        return True

    def _sum_prices(match=None):
        return sum(float(p.get('price', 0)) for p in daily_prices if _in_range(p.get('date', '')) and (not match or match(p)))

    def _sum_extra(match=None):
        return sum(float(e.get('amount', 0)) for e in extra_incomes if _in_range(e.get('date', '')) and (not match or match(e)))

    def _sum_expenses(match=None):
        return sum(float(e.get('amount', 0)) for e in expenses if _in_range(e.get('date', '')) and (not match or match(e)))

    def _days_in_month(y, m):
        return [31,28+(1 if y%4==0 and (y%100!=0 or y%400==0) else 0),31,30,31,30,31,31,30,31,30,31][m-1]

    # 自动确定时间范围
    now = datetime.now()
    if not df:
        df = datetime(now.year, 1, 1) if mode == 'yearly' else datetime(now.year, now.month, 1)
    if not dt and mode == 'yearly':
        dt = datetime(now.year, 12, 31)
    elif not dt and mode == 'monthly':
        dt = datetime(now.year, now.month, _days_in_month(now.year, now.month))

    labels, income_data, expense_data, occupancy_data = [], [], [], []

    if mode in ('yearly', 'monthly'):
        cur = datetime(df.year, df.month, 1)
        end = dt
        step = timedelta(days=365) if mode == 'yearly' else timedelta(days=32)
        while cur <= end:
            if mode == 'yearly':
                labels.append(str(cur.year))
                ym = str(cur.year)
                income_data.append(round(_sum_prices(lambda p: p['date'].startswith(ym)) + _sum_extra(lambda e: e['date'].startswith(ym)), 2))
                expense_data.append(round(_sum_expenses(lambda e: e['date'].startswith(ym)), 2))
                priced_days = sum(1 for p in daily_prices if p['date'].startswith(ym) and p.get('price',0)>0)
                total = room_count * (366 if cur.year%4==0 else 365) or 1
                occupancy_data.append(round(priced_days/total*100,1))
                cur = datetime(cur.year+1, 1, 1)
            else:
                labels.append(f'{cur.year}-{cur.month:02d}')
                ym = cur.strftime('%Y-%m')
                income_data.append(round(_sum_prices(lambda p: p['date'].startswith(ym)) + _sum_extra(lambda e: e['date'].startswith(ym)), 2))
                expense_data.append(round(_sum_expenses(lambda e: e['date'].startswith(ym)), 2))
                dim = _days_in_month(cur.year, cur.month)
                priced_days = sum(1 for p in daily_prices if p['date'].startswith(ym) and p.get('price',0)>0)
                total = room_count * dim or 1
                occupancy_data.append(round(priced_days/total*100,1))
                cur = datetime(cur.year + (cur.month//12), (cur.month%12)+1, 1)

    elif mode == 'weekly':
        if not df: df = datetime(now.year, now.month, 1)
        dim = _days_in_month(df.year, df.month)
        end = dt or datetime(df.year, df.month, dim)
        cur = df
        wn = 1
        while cur <= end:
            we = min(cur + timedelta(days=6), end)
            labels.append(f'W{wn}')
            wn += 1
            income_data.append(round(_sum_prices(lambda p: cur.strftime('%Y-%m-%d') <= p['date'] <= we.strftime('%Y-%m-%d')) + _sum_extra(lambda e: cur.strftime('%Y-%m-%d') <= e['date'] <= we.strftime('%Y-%m-%d')), 2))
            expense_data.append(round(_sum_expenses(lambda e: cur.strftime('%Y-%m-%d') <= e['date'] <= we.strftime('%Y-%m-%d')), 2))
            pd_cnt, wd = 0, 0
            d = cur
            while d <= we:
                ds = d.strftime('%Y-%m-%d')
                pd_cnt += sum(1 for p in daily_prices if p['date']==ds and p.get('price',0)>0)
                wd += 1; d += timedelta(days=1)
            occupancy_data.append(round(pd_cnt/(room_count*wd)*100,1) if room_count*wd else 0)
            cur += timedelta(days=7)

    elif mode == 'rooms':
        for room in rooms:
            labels.append(room.get('room_number', f'R{room["id"]}'))
            rm = str(room['id'])
            income_data.append(round(sum(float(p.get('price',0)) for p in daily_prices if str(p.get('room_id'))==rm and _in_range(p.get('date',''))), 2))
            expense_data.append(round(sum(float(e.get('amount',0)) for e in expenses if str(e.get('room_id'))==rm and _in_range(e.get('date',''))), 2))
            pd_cnt = sum(1 for p in daily_prices if str(p.get('room_id'))==rm and _in_range(p.get('date','')) and p.get('price',0)>0)
            total_days = (dt - df).days + 1 if df and dt else 30
            occupancy_data.append(round(pd_cnt/total_days*100,1) if total_days else 0)

    elif mode == 'daily':
        if df and dt:
            cur = df
            while cur <= dt:
                ds = cur.strftime('%Y-%m-%d')
                labels.append(f'{cur.month}/{cur.day}')
                income_data.append(round(_sum_prices(lambda p: p['date']==ds) + _sum_extra(lambda e: e['date']==ds), 2))
                expense_data.append(round(_sum_expenses(lambda e: e['date']==ds), 2))
                occupancy_data.append(0)
                cur += timedelta(days=1)

    # 总计
    total_income = sum(income_data)
    total_expense = sum(expense_data)
    total_profit = total_income - total_expense
    avg_occupancy = round(sum(occupancy_data)/len(occupancy_data),1) if occupancy_data else 0

    return jsonify({'success': True, 'data': {
        'mode': mode, 'labels': labels, 'income': income_data, 'expense': expense_data,
        'occupancy': occupancy_data, 'rooms': rooms,
        'total_income': round(total_income,2), 'total_expense': round(total_expense,2),
        'total_profit': round(total_profit,2), 'avg_occupancy': avg_occupancy,
        'date_from': df.strftime('%Y-%m-%d') if df else '', 'date_to': dt.strftime('%Y-%m-%d') if dt else ''
    }})


# ==================== API：日历数据 ====================

@app.route('/api/calendar-data')
def api_calendar_data():
    """获取日历视图数据"""
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))

    bookings = DataManager.read_json(BOOKINGS_FILE, [])
    rooms = DataManager.read_json(ROOMS_FILE, [])
    rooms.sort(key=lambda r: r.get('sort_order', 0))
    guests = DataManager.read_json(GUESTS_FILE, [])

    guest_map = {g['id']: g for g in guests}

    # 预订来源颜色
    sources = DataManager.read_json(BOOKING_SOURCES_FILE, [])
    source_colors = {s['name']: s['color'] for s in sources}

    # 筛选该月的预订
    month_start = f'{year}-{month:02d}-01'
    if month == 12:
        month_end = f'{year + 1}-01-01'
    else:
        month_end = f'{year}-{month + 1:02d}-01'

    month_bookings = [
        b for b in bookings
        if b.get('status') not in ('cancelled', 'no_show') and
           b.get('check_in_date', '') < month_end and
           b.get('check_out_date', '') > month_start
    ]

    events = []
    for b in month_bookings:
        guest = guest_map.get(b.get('guest_id'), {})
        src = b.get('source', '')
        events.append({
            'id': b['id'],
            'room_id': b['room_id'],
            'title': f"{guest.get('name', '未知')}",
            'source': src,
            'start': b['check_in_date'],
            'end': b['check_out_date'],
            'status': b['status'],
            'color': source_colors.get(src, '#6b7280')
        })

    # 筛选该月的每日价格
    all_prices = DataManager.read_json(DAILY_PRICES_FILE, [])
    month_prices = {}
    for p in all_prices:
        if p.get('date', '').startswith(f'{year}-{month:02d}'):
            key = f"{p.get('room_id', 0)}_{p.get('date', '')}"
            month_prices[key] = p

    # 节假日
    holidays = DataManager.read_json(HOLIDAYS_FILE, [])
    month_holidays = {}
    for h in holidays:
        h_date = h.get('date', '')
        if h_date.startswith(f'{year}-{month:02d}'):
            month_holidays[h_date] = {'name': h.get('name', ''), 'type': 'holiday'}
            # 前一天标记为节前日
            h_dt = parse_date(h_date)
            if h_dt:
                eve = (h_dt - timedelta(days=1)).strftime('%Y-%m-%d')
                if eve.startswith(f'{year}-{month:02d}') and eve not in month_holidays:
                    month_holidays[eve] = {'name': h.get('name', '') + '前', 'type': 'holiday_eve'}

    return jsonify({
        'success': True,
        'data': {
            'year': year,
            'month': month,
            'rooms': rooms,
            'events': events,
            'prices': month_prices,
            'sources': sources,
            'holidays': month_holidays
        }
    })


@app.route('/api/calendar/export-excel')
def api_calendar_export_excel():
    """导出日历为 Excel"""
    year = int(request.args.get('year', datetime.now().year))
    month = int(request.args.get('month', datetime.now().month))

    rooms = DataManager.read_json(ROOMS_FILE, [])
    prices = DataManager.read_json(DAILY_PRICES_FILE, [])
    bookings = DataManager.read_json(BOOKINGS_FILE, [])
    guests = DataManager.read_json(GUESTS_FILE, [])
    guest_map = {g['id']: g.get('name', '') for g in guests}

    # 价格索引
    price_map = {}
    for p in prices:
        price_map[(p.get('room_id'), p.get('date'))] = p.get('price', 0)

    # 当月天数
    days_in_month = [31, 28 + (1 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 0),
                    31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1]

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = f'{year}年{month}月'

    # 标题行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=days_in_month + 2)
    ws.cell(1, 1, f'{year}年{month}月 - 民宿房间日历').font = Font(size=14, bold=True)
    ws.cell(1, 1).alignment = Alignment(horizontal='center')

    # 表头
    headers = ['房间号', '房型']
    for d in range(1, days_in_month + 1):
        headers.append(f'{month}/{d}')
    headers.append('月合计')
    for i, h in enumerate(headers, 1):
        cell = ws.cell(3, i, h)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill('solid', fgColor='E5E7EB')
        cell.alignment = Alignment(horizontal='center')

    # 数据行
    for ri, room in enumerate(rooms):
        row = ri + 4
        ws.cell(row, 1, room.get('room_number', ''))
        type_map = {t['id']: t['name'] for t in DataManager.read_json(ROOM_TYPES_FILE, [])}
        ws.cell(row, 2, type_map.get(room.get('room_type_id'), ''))

        month_total = 0
        for d in range(1, days_in_month + 1):
            date_str = f'{year}-{month:02d}-{d:02d}'
            price = price_map.get((room.get('id'), date_str), 0)
            if price > 0:
                ws.cell(row, d + 2, price)
                month_total += price
            # 预订人标注（用备注）
            day_bookings = [b for b in bookings
                          if b.get('room_id') == room.get('id') and
                          b.get('check_in_date', '') <= date_str and
                          b.get('check_out_date', '') > date_str and
                          b.get('status') not in ('cancelled', 'no_show')]
            if day_bookings:
                guest_name = guest_map.get(day_bookings[0].get('guest_id'), '')
                cell = ws.cell(row, d + 2)
                if guest_name:
                    cell.comment = None  # openpyxl comment can be complex, skip for now

        ws.cell(row, days_in_month + 3, round(month_total, 2))
        ws.cell(row, days_in_month + 3).font = Font(bold=True)

    # 合计行
    total_row = len(rooms) + 4
    ws.cell(total_row, 1, '日合计').font = Font(bold=True)
    for d in range(1, days_in_month + 1):
        col_total = 0
        for ri in range(len(rooms)):
            date_str = f'{year}-{month:02d}-{d:02d}'
            price = price_map.get((rooms[ri].get('id'), date_str), 0)
            col_total += price
        cell = ws.cell(total_row, d + 2, round(col_total, 2))
        cell.font = Font(bold=True)

    # 调整列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    for d in range(1, days_in_month + 1):
        ws.column_dimensions[ws.cell(3, d + 2).column_letter].width = 8

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'calendar_{year}_{month:02d}.xlsx')


def _get_status_color(status):
    colors = {
        'pending': '#f59e0b',      # 黄色
        'confirmed': '#3b82f6',    # 蓝色
        'checked_in': '#10b981',   # 绿色
        'checked_out': '#6b7280',  # 灰色
    }
    return colors.get(status, '#6b7280')


def _auto_fill_daily_prices(room_id, check_in, check_out, total_price, force=False):
    """预订时自动将日均价填入日历日期范围"""
    try:
        start = parse_date(check_in)
        end = parse_date(check_out)
        if not start or not end:
            return
        days = (end - start).days
        if days <= 0:
            return
        daily = round(total_price / days, 2)

        prices = DataManager.read_json(DAILY_PRICES_FILE, [])
        existing_map = {}
        for p in prices:
            existing_map[(p.get('room_id'), p.get('date'))] = p

        current = start
        while current < end:
            date_str = current.strftime('%Y-%m-%d')
            key = (room_id, date_str)
            if key in existing_map:
                if force:
                    # 强制覆盖已有价格
                    existing_map[key]['price'] = daily
                    existing_map[key]['note'] = '预订覆盖'
                    existing_map[key]['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            else:
                new_p = {
                    'id': DataManager.generate_id(prices),
                    'room_id': room_id,
                    'date': date_str,
                    'price': daily,
                    'note': '自动填价',
                    'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                prices.append(new_p)
                existing_map[key] = new_p
            current += timedelta(days=1)

        DataManager.write_json(DAILY_PRICES_FILE, prices)
    except Exception:
        pass  # 自动填价失败不影响预订


def _delete_daily_prices_range(room_id, check_in, check_out):
    """删除指定日期范围内的每日价格"""
    try:
        start = parse_date(check_in)
        end = parse_date(check_out)
        if not start or not end:
            return
        prices = DataManager.read_json(DAILY_PRICES_FILE, [])
        to_delete = set()
        current = start
        while current < end:
            date_str = current.strftime('%Y-%m-%d')
            for i, p in enumerate(prices):
                if p.get('room_id') == room_id and p.get('date') == date_str:
                    to_delete.add(i)
            current += timedelta(days=1)
        if to_delete:
            prices = [p for i, p in enumerate(prices) if i not in to_delete]
            DataManager.write_json(DAILY_PRICES_FILE, prices)
            SystemLogger.info('清除每日价格',
                              f'房间{room_id} {check_in}~{check_out}, 共{len(to_delete)}天')
    except Exception:
        pass


# ==================== API：图片上传 ====================

@app.route('/api/upload', methods=['POST'])
def api_upload():
    """上传图片"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '没有选择文件'}), 400

    file = request.files['file']
    subfolder = request.form.get('subfolder', '')

    result = ImageHandler.save_image(file, subfolder)
    if result['success']:
        SystemLogger.info('上传图片', f'文件名: {result["filename"]}')
        return jsonify(result)
    return jsonify(result), 400


@app.route('/api/upload/<subfolder>', methods=['POST'])
def api_upload_subfolder(subfolder):
    """上传图片到指定子目录"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': '没有选择文件'}), 400

    file = request.files['file']
    result = ImageHandler.save_image(file, subfolder)
    if result['success']:
        SystemLogger.info('上传图片',
                          f'目录: {subfolder}, 文件: {result["filename"]}')
        return jsonify(result)
    return jsonify(result), 400


@app.route('/api/delete-image', methods=['POST'])
def api_delete_image():
    """删除图片"""
    data = request.get_json()
    filename = data.get('filename', '')
    subfolder = data.get('subfolder', '')

    if not filename:
        return jsonify({'success': False, 'error': '未指定文件名'}), 400

    if ImageHandler.delete_image(filename, subfolder):
        SystemLogger.info('删除图片', f'文件名: {filename}')
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': '删除失败'}), 400


# ==================== API：日志管理 ====================

@app.route('/api/logs')
def api_logs():
    """获取系统日志"""
    level = request.args.get('level', '')
    page = int(request.args.get('page', 1))
    result = SystemLogger.get_logs(level=level if level else None,
                                   page=page, page_size=50)
    return jsonify({'success': True, 'data': result})


@app.route('/api/logs/clear', methods=['POST'])
def api_logs_clear():
    """清空日志"""
    SystemLogger.clear_logs()
    return jsonify({'success': True})


# ==================== API：数据导出 ====================

@app.route('/api/export')
def api_export():
    """导出所有数据（用于数据迁移）"""
    data_type = request.args.get('type', 'all')

    export_data = {}
    if data_type in ('all', 'room_types'):
        export_data['room_types'] = DataManager.read_json(ROOM_TYPES_FILE, [])
    if data_type in ('all', 'rooms'):
        export_data['rooms'] = DataManager.read_json(ROOMS_FILE, [])
    if data_type in ('all', 'guests'):
        export_data['guests'] = DataManager.read_json(GUESTS_FILE, [])
    if data_type in ('all', 'bookings'):
        export_data['bookings'] = DataManager.read_json(BOOKINGS_FILE, [])
    if data_type in ('all', 'logs'):
        export_data['logs'] = DataManager.read_json(LOGS_FILE, [])
    if data_type in ('all', 'special_prices'):
        export_data['daily_prices'] = DataManager.read_json(DAILY_PRICES_FILE, [])
    if data_type in ('all', 'extra_income'):
        export_data['extra_income'] = DataManager.read_json(EXTRA_INCOME_FILE, [])
    if data_type in ('all', 'diary'):
        export_data['diary'] = DataManager.read_json(DIARY_FILE, [])

    export_data['export_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    export_data['version'] = '1.0'

    return jsonify({'success': True, 'data': export_data})


@app.route('/api/import', methods=['POST'])
def api_import():
    """导入数据"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': '没有数据'}), 400

        count = 0
        for key, filepath in [
            ('room_types', ROOM_TYPES_FILE),
            ('rooms', ROOMS_FILE),
            ('guests', GUESTS_FILE),
            ('bookings', BOOKINGS_FILE),
            ('special_prices', DAILY_PRICES_FILE),
            ('daily_prices', DAILY_PRICES_FILE),
            ('extra_income', EXTRA_INCOME_FILE),
            ('diary', DIARY_FILE),
        ]:
            if key in data and isinstance(data[key], list):
                DataManager.write_json(filepath, data[key])
                count += len(data[key])

        SystemLogger.info('导入数据', f'共导入 {count} 条记录')
        return jsonify({'success': True, 'message': f'成功导入 {count} 条记录'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


# ==================== API：每日价格管理（日历点击设置） ====================

@app.route('/api/daily-prices', methods=['GET', 'POST'])
def api_daily_prices():
    if request.method == 'GET':
        prices = DataManager.read_json(DAILY_PRICES_FILE, [])
        room_id = request.args.get('room_id', '')
        date = request.args.get('date', '')
        if room_id:
            prices = [p for p in prices if str(p.get('room_id')) == room_id]
        if date:
            prices = [p for p in prices if p.get('date') == date]
        return jsonify({'success': True, 'data': prices})

    elif request.method == 'POST':
        try:
            data = request.get_json()
            prices = DataManager.read_json(DAILY_PRICES_FILE, [])
            room_id = int(data.get('room_id', 0))
            date = data.get('date', '').strip()
            price_val = float(data.get('price', 0))

            # 价格为 0 或负数 → 删除记录（清除价格）
            if price_val <= 0:
                existing = [p for p in prices
                            if p.get('room_id') == room_id and p.get('date') == date]
                if existing:
                    DataManager.delete_by_id(prices, existing[0]['id'])
                    DataManager.write_json(DAILY_PRICES_FILE, prices)
                    SystemLogger.info('清除每日价格', f'房间{room_id} {date}')
                return jsonify({'success': True, 'data': None, 'cleared': True})

            # 检查是否已存在，直接覆盖
            existing = [p for p in prices
                        if p.get('room_id') == room_id and p.get('date') == date]
            if existing:
                existing[0]['price'] = price_val
                existing[0]['note'] = data.get('note', '').strip()
                existing[0]['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                DataManager.write_json(DAILY_PRICES_FILE, prices)
                SystemLogger.info('更新每日价格',
                                  f'房间{room_id} {date}: ¥{price_val}')
                return jsonify({'success': True, 'data': existing[0], 'updated': True})

            new_price = {
                'id': DataManager.generate_id(prices),
                'room_id': room_id,
                'date': date,
                'price': price_val,
                'note': data.get('note', '').strip(),
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            prices.append(new_price)
            DataManager.write_json(DAILY_PRICES_FILE, prices)
            SystemLogger.info('设置每日价格',
                              f'房间{room_id} {date}: ¥{price_val}')
            return jsonify({'success': True, 'data': new_price})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/daily-prices/<int:price_id>', methods=['DELETE'])
def api_daily_price_delete(price_id):
    prices = DataManager.read_json(DAILY_PRICES_FILE, [])
    item = DataManager.get_by_id(prices, price_id)
    if DataManager.delete_by_id(prices, price_id):
        DataManager.write_json(DAILY_PRICES_FILE, prices)
        SystemLogger.info('删除每日价格',
                          f'房间{item.get("room_id")} {item.get("date")}')
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': '记录不存在'}), 404


# ==================== API：预订来源管理 ====================

@app.route('/api/booking-sources', methods=['GET', 'POST'])
def api_booking_sources():
    if request.method == 'GET':
        sources = DataManager.read_json(BOOKING_SOURCES_FILE, [])
        return jsonify({'success': True, 'data': sources})

    elif request.method == 'POST':
        try:
            data = request.get_json()
            sources = DataManager.read_json(BOOKING_SOURCES_FILE, [])
            new_src = {
                'id': DataManager.generate_id(sources),
                'name': data.get('name', '').strip(),
                'color': data.get('color', '#6b7280'),
            }
            sources.append(new_src)
            DataManager.write_json(BOOKING_SOURCES_FILE, sources)
            SystemLogger.info('新增预订来源', f'{new_src["name"]}')
            return jsonify({'success': True, 'data': new_src})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/booking-sources/<int:src_id>', methods=['PUT', 'DELETE'])
def api_booking_source_detail(src_id):
    sources = DataManager.read_json(BOOKING_SOURCES_FILE, [])
    if request.method == 'PUT':
        try:
            data = request.get_json()
            updates = {
                'name': data.get('name', '').strip(),
                'color': data.get('color', '#6b7280'),
            }
            if DataManager.update_by_id(sources, src_id, updates):
                DataManager.write_json(BOOKING_SOURCES_FILE, sources)
                return jsonify({'success': True})
            return jsonify({'success': False, 'error': '不存在'}), 404
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400
    elif request.method == 'DELETE':
        if DataManager.delete_by_id(sources, src_id):
            DataManager.write_json(BOOKING_SOURCES_FILE, sources)
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '不存在'}), 404


# ==================== API：节假日管理 ====================

@app.route('/api/holidays', methods=['GET', 'POST'])
def api_holidays():
    if request.method == 'GET':
        holidays = DataManager.read_json(HOLIDAYS_FILE, [])
        holidays.sort(key=lambda x: x.get('date', ''))
        return jsonify({'success': True, 'data': holidays})

    elif request.method == 'POST':
        try:
            data = request.get_json()
            holidays = DataManager.read_json(HOLIDAYS_FILE, [])
            new_h = {
                'id': DataManager.generate_id(holidays),
                'date': data.get('date', '').strip(),
                'name': data.get('name', '').strip(),
            }
            holidays.append(new_h)
            DataManager.write_json(HOLIDAYS_FILE, holidays)
            return jsonify({'success': True, 'data': new_h})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/holidays/<int:h_id>', methods=['DELETE'])
def api_holiday_delete(h_id):
    holidays = DataManager.read_json(HOLIDAYS_FILE, [])
    if DataManager.delete_by_id(holidays, h_id):
        DataManager.write_json(HOLIDAYS_FILE, holidays)
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': '不存在'}), 404


# ==================== API：支出类别管理 ====================

@app.route('/api/expense-categories', methods=['GET', 'POST'])
def api_expense_categories():
    if request.method == 'GET':
        cats = DataManager.read_json(EXPENSE_CATEGORIES_FILE, [])
        return jsonify({'success': True, 'data': cats})
    elif request.method == 'POST':
        try:
            data = request.get_json()
            cats = DataManager.read_json(EXPENSE_CATEGORIES_FILE, [])
            new_c = {'id': DataManager.generate_id(cats), 'name': data.get('name', '').strip(), 'color': data.get('color', '#6b7280')}
            cats.append(new_c)
            DataManager.write_json(EXPENSE_CATEGORIES_FILE, cats)
            return jsonify({'success': True, 'data': new_c})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/expense-categories/<int:c_id>', methods=['PUT', 'DELETE'])
def api_expense_category_detail(c_id):
    cats = DataManager.read_json(EXPENSE_CATEGORIES_FILE, [])
    if request.method == 'PUT':
        data = request.get_json()
        if DataManager.update_by_id(cats, c_id, {'name': data.get('name', '').strip(), 'color': data.get('color', '#6b7280')}):
            DataManager.write_json(EXPENSE_CATEGORIES_FILE, cats)
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '不存在'}), 404
    elif request.method == 'DELETE':
        if DataManager.delete_by_id(cats, c_id):
            DataManager.write_json(EXPENSE_CATEGORIES_FILE, cats)
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '不存在'}), 404


# ==================== API：支出管理 ====================

@app.route('/api/expenses', methods=['GET', 'POST'])
def api_expenses():
    if request.method == 'GET':
        page = int(request.args.get('page', 1))
        keyword = request.args.get('keyword', '')
        room_id = request.args.get('room_id', '')
        date_from = request.args.get('date_from', '')
        date_to = request.args.get('date_to', '')
        expenses = DataManager.read_json(EXPENSES_FILE, [])
        if room_id:
            expenses = [e for e in expenses if str(e.get('room_id')) == room_id]
        if date_from:
            expenses = [e for e in expenses if e.get('date', '') >= date_from]
        if date_to:
            expenses = [e for e in expenses if e.get('date', '') <= date_to]
        if keyword:
            expenses = [e for e in expenses if keyword.lower() in e.get('note', '').lower() or keyword.lower() in e.get('category', '').lower()]
        result = DataManager.query(expenses, sort_key='date', sort_reverse=True, page=page, page_size=PAGE_SIZE)
        return jsonify({'success': True, 'data': result})
    elif request.method == 'POST':
        try:
            data = request.get_json()
            expenses = DataManager.read_json(EXPENSES_FILE, [])
            new_e = {
                'id': DataManager.generate_id(expenses), 'date': data.get('date', get_today_str()),
                'amount': float(data.get('amount', 0)), 'category': data.get('category', '').strip(),
                'room_id': int(data.get('room_id', 0)) or None, 'note': data.get('note', '').strip(),
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            expenses.append(new_e)
            DataManager.write_json(EXPENSES_FILE, expenses)
            SystemLogger.info('新增支出', f'{new_e["date"]}: ¥{new_e["amount"]} - {new_e["category"]}')
            return jsonify({'success': True, 'data': new_e})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/expenses/<int:e_id>', methods=['PUT', 'DELETE'])
def api_expense_detail(e_id):
    expenses = DataManager.read_json(EXPENSES_FILE, [])
    if request.method == 'PUT':
        data = request.get_json()
        updates = {'date': data.get('date', ''), 'amount': float(data.get('amount', 0)), 'category': data.get('category', ''), 'room_id': int(data.get('room_id', 0)) or None, 'note': data.get('note', '')}
        if DataManager.update_by_id(expenses, e_id, updates):
            DataManager.write_json(EXPENSES_FILE, expenses)
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '不存在'}), 404
    elif request.method == 'DELETE':
        if DataManager.delete_by_id(expenses, e_id):
            DataManager.write_json(EXPENSES_FILE, expenses)
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '不存在'}), 404


# ==================== API：额外收入管理 ====================

@app.route('/api/extra-income', methods=['GET', 'POST'])
def api_extra_income():
    if request.method == 'GET':
        page = int(request.args.get('page', 1))
        date = request.args.get('date', '')
        incomes = DataManager.read_json(EXTRA_INCOME_FILE, [])
        if date:
            incomes = [e for e in incomes if e.get('date') == date]
        result = DataManager.query(incomes, sort_key='date', sort_reverse=True,
                                   page=page, page_size=PAGE_SIZE)
        return jsonify({'success': True, 'data': result})

    elif request.method == 'POST':
        try:
            data = request.get_json()
            incomes = DataManager.read_json(EXTRA_INCOME_FILE, [])
            new_entry = {
                'id': DataManager.generate_id(incomes),
                'date': data.get('date', get_today_str()),
                'amount': float(data.get('amount', 0)),
                'description': data.get('description', '').strip(),
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            incomes.append(new_entry)
            DataManager.write_json(EXTRA_INCOME_FILE, incomes)
            SystemLogger.info('新增额外收入',
                              f'{new_entry["date"]}: ¥{new_entry["amount"]} - {new_entry["description"]}')
            return jsonify({'success': True, 'data': new_entry})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/extra-income/<int:income_id>', methods=['PUT', 'DELETE'])
def api_extra_income_detail(income_id):
    incomes = DataManager.read_json(EXTRA_INCOME_FILE, [])

    if request.method == 'PUT':
        try:
            data = request.get_json()
            updates = {
                'date': data.get('date', get_today_str()),
                'amount': float(data.get('amount', 0)),
                'description': data.get('description', '').strip(),
            }
            if DataManager.update_by_id(incomes, income_id, updates):
                DataManager.write_json(EXTRA_INCOME_FILE, incomes)
                return jsonify({'success': True})
            return jsonify({'success': False, 'error': '记录不存在'}), 404
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400

    elif request.method == 'DELETE':
        if DataManager.delete_by_id(incomes, income_id):
            DataManager.write_json(EXTRA_INCOME_FILE, incomes)
            SystemLogger.info('删除额外收入', f'ID: {income_id}')
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '记录不存在'}), 404


# ==================== API：运营日记 ====================

@app.route('/api/diary', methods=['GET', 'POST'])
def api_diary():
    if request.method == 'GET':
        page = int(request.args.get('page', 1))
        keyword = request.args.get('keyword', '')
        diary_list = DataManager.read_json(DIARY_FILE, [])

        if keyword:
            diary_list = DataManager.search(diary_list, keyword,
                                            ['title', 'content', 'weather', 'mood'])

        result = DataManager.query(diary_list, sort_key='date',
                                   sort_reverse=True, page=page,
                                   page_size=PAGE_SIZE)
        return jsonify({'success': True, 'data': result})

    elif request.method == 'POST':
        try:
            data = request.get_json()
            diary_list = DataManager.read_json(DIARY_FILE, [])

            new_entry = {
                'id': DataManager.generate_id(diary_list),
                'date': data.get('date', get_today_str()),
                'title': data.get('title', '').strip(),
                'content': data.get('content', '').strip(),
                'weather': data.get('weather', ''),
                'mood': data.get('mood', ''),
                'images': data.get('images', []),
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            diary_list.append(new_entry)
            DataManager.write_json(DIARY_FILE, diary_list)
            SystemLogger.info('新增日记', f'日期: {new_entry["date"]}, 标题: {new_entry["title"]}')
            return jsonify({'success': True, 'data': new_entry})
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/diary/<int:entry_id>', methods=['GET', 'PUT', 'DELETE'])
def api_diary_detail(entry_id):
    diary_list = DataManager.read_json(DIARY_FILE, [])

    if request.method == 'GET':
        item = DataManager.get_by_id(diary_list, entry_id)
        if item:
            return jsonify({'success': True, 'data': item})
        return jsonify({'success': False, 'error': '日记不存在'}), 404

    elif request.method == 'PUT':
        try:
            data = request.get_json()
            updates = {
                'date': data.get('date', get_today_str()),
                'title': data.get('title', '').strip(),
                'content': data.get('content', '').strip(),
                'weather': data.get('weather', ''),
                'mood': data.get('mood', ''),
                'images': data.get('images', []),
            }
            if DataManager.update_by_id(diary_list, entry_id, updates):
                DataManager.write_json(DIARY_FILE, diary_list)
                SystemLogger.info('更新日记', f'日记ID: {entry_id}')
                return jsonify({'success': True})
            return jsonify({'success': False, 'error': '日记不存在'}), 404
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 400

    elif request.method == 'DELETE':
        if DataManager.delete_by_id(diary_list, entry_id):
            DataManager.write_json(DIARY_FILE, diary_list)
            SystemLogger.info('删除日记', f'日记ID: {entry_id}')
            return jsonify({'success': True})
        return jsonify({'success': False, 'error': '日记不存在'}), 404


# ==================== API：批量更新房间状态 ====================

@app.route('/api/rooms/batch-status', methods=['POST'])
def api_batch_room_status():
    """批量更新房间状态"""
    try:
        data = request.get_json()
        room_ids = data.get('room_ids', [])
        new_status = data.get('status', '')

        if not room_ids or new_status not in ROOM_STATUS:
            return jsonify({'success': False, 'error': '参数错误'}), 400

        rooms = DataManager.read_json(ROOMS_FILE, [])
        count = 0
        for rid in room_ids:
            if DataManager.update_by_id(rooms, rid, {'status': new_status}):
                count += 1
        DataManager.write_json(ROOMS_FILE, rooms)

        SystemLogger.info('批量更新房间状态',
                          f'更新 {count} 个房间状态为: {ROOM_STATUS.get(new_status)}')
        return jsonify({'success': True, 'message': f'成功更新 {count} 个房间'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


@app.route('/api/rooms/sort', methods=['POST'])
def api_rooms_sort():
    """更新房间排序"""
    try:
        data = request.get_json()
        room_ids = data.get('room_ids', [])
        rooms = DataManager.read_json(ROOMS_FILE, [])
        for i, rid in enumerate(room_ids):
            DataManager.update_by_id(rooms, int(rid), {'sort_order': i})
        DataManager.write_json(ROOMS_FILE, rooms)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


# ==================== API：清除所有数据 ====================

@app.route('/api/clear-all-data', methods=['POST'])
def api_clear_all_data():
    """清除所有业务数据"""
    try:
        for f in [ROOM_TYPES_FILE, ROOMS_FILE, GUESTS_FILE, BOOKINGS_FILE,
                  DAILY_PRICES_FILE, EXTRA_INCOME_FILE, DIARY_FILE, LOGS_FILE]:
            DataManager.write_json(f, [])
        SystemLogger.info('清除所有数据', '所有业务数据已清空')
        return jsonify({'success': True, 'message': '所有数据已清除'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== 错误处理 ====================

@app.errorhandler(404)
def not_found(e):
    if request.path.startswith('/api/'):
        return jsonify({'success': False, 'error': '接口不存在'}), 404
    return render_template('index.html'), 404


@app.errorhandler(500)
def server_error(e):
    SystemLogger.error('服务器错误', str(e))
    if request.path.startswith('/api/'):
        return jsonify({'success': False, 'error': '服务器内部错误'}), 500
    return render_template('index.html'), 500


# ==================== 启动应用 ====================

if __name__ == '__main__':
    print('=' * 50)
    print(f'  {HOMESTAY_NAME}')
    print(f'  访问地址: http://0.0.0.0:{PORT}')
    print(f'  数据目录: {DATA_DIR}')
    print(f'  上传目录: {UPLOAD_DIR}')
    print(f'  日志目录: {LOG_DIR}')
    print('=' * 50)

    # 记录启动日志
    SystemLogger.info('系统启动', f'端口: {PORT}')

    app.run(host=HOST, port=PORT, debug=DEBUG, threaded=True)
